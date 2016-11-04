from django.shortcuts import render_to_response , render
from django.template import RequestContext
from planilla.apps.home.forms import *
from planilla.apps.home.models import *
from django.contrib.auth import *
from django.contrib.auth.models import *
from planilla.apps.home.forms import *
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, InvalidPage
from django.shortcuts import render_to_response 
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django.core.mail import send_mail
from django.forms import ModelForm
from planilla.apps.home.forms import estado_contrato_form
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from django.http import HttpResponse
from openpyxl import load_workbook


def mostrarTabla(request):
	contratistas = Contratista.objects.all()
	return render_to_response('home/tabla.html',{'contratistas': contratistas})	


#from django.views.generic import TemplateView
from django.http import HttpResponseRedirect

def index_view(request):
	return render_to_response('home/index.html',context_instance = RequestContext(request))

def acerca_view(request):
	return render_to_response('home/acerca.html',context_instance = RequestContext(request))

def registrocontrato_view (request):

	return render_to_response('home/registrocontrato.html',context_instance = RequestContext(request))

def login_view(request):
	mensaje = ''
	if request.user.is_authenticated():
			current_user = request.user
			request.session['name'] = request.user.id
			print request.session['name']
			ctx = { 'mensaje':mensaje}
			return HttpResponseRedirect('/', ctx)

	elif request.user.is_staff:
			current_user = request.user
			request.session['name'] = request.user.id
			ctx = { 'mensaje':mensaje}
			#return HttpResponseRedirect('/gestion/', ctx)
	else:
		usu=''
		pas=''
		formulario = Login_form(request.POST)
		if request.method == "POST":
			
			if formulario.is_valid():
				usu = formulario.cleaned_data['usuario']
				pas = formulario.cleaned_data['clave']
				usuario = authenticate(username = usu, password = pas)
				if usuario is not None and usuario.is_staff:
					current_user = request.user
					request.session['name'] = request.user.id
					print request.session['name']
					login(request, usuario)
					return HttpResponseRedirect('/gestion/')
				if usuario is not None and usuario.is_active:
					login(request, usuario)
					return HttpResponseRedirect('/')
				else:
					mensaje = "usuario y/o clave incorrecta: "

					
	formulario = Login_form()
	ctx = {'form':formulario, 'mensaje':mensaje}
	return render_to_response('home/login.html', ctx, context_instance = RequestContext(request))

def logout_view(request):
	logout(request)
	return HttpResponseRedirect('/')

def registrarContratista_view(request):
	form = registrarAdminForm()
	if request.method == "POST":
		form = registrarAdminForm(request.POST)
		if form.is_valid():
			usuario = form.cleaned_data['username']
			email = form.cleaned_data['email']
			password_one = form.cleaned_data['password_one']
			password_two = form.cleaned_data['password_two']
			u = User.objects.create_user(username=usuario,email=email,password=password_one)
			u.save()
			return render_to_response('home/Successful registration.html',context_instance=RequestContext(request))
		else:
			ctx = {'form':form}
			return render_to_response('home/registrarusuario.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form}
	return render_to_response('home/registrarusuario.html',ctx,context_instance=RequestContext(request))

def registrarsupervisor_view(request):
	form = registrarSupervisorForm()
	if request.method == "POST":
		form = registrarSupervisorForm(request.POST)
		if form.is_valid():
			nombre = form.cleaned_data['nombre']
			apellido = form.cleaned_data['apellido']
			usuario = form.cleaned_data['username']
			email = form.cleaned_data['email']
			password_one = form.cleaned_data['password_one']
			password_two = form.cleaned_data['password_two']

			u = User.objects.create_user(username=usuario,email=email,password=password_one,first_name=nombre,last_name=apellido)
			u.is_staff=True
			u.save()
			datos = Supervisor.objects.create(nombre = nombre, apellido= apellido, correo= email)
			datos.save


			return render_to_response('home/Successful registration.html',context_instance=RequestContext(request))
		else:
			ctx = {'form':form}
			return render_to_response('home/registrarsupervisor.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form}
	return render_to_response('home/registrarsupervisor.html',ctx,context_instance=RequestContext(request))

def supervisor_view (request):
	if request.user.is_staff:
		return render_to_response('home/supervisor.html',context_instance = RequestContext(request))
	else:
		return render_to_response('home/index.html',context_instance = RequestContext(request))


def register_view(request):
	form = Register_Form(request.POST or None)
	if request.method == "POST":
			
		if form.is_valid():
			instance = form.save(commit=False)
			instance.save()
			
			context = {
				"form": form
			} 
			return render(request, 'home/Successful registration.html',locals() )
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/register.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/register.html',locals() )
	

def datoscontratista_view(request):
	form = DatosContratista_Form(request.POST or None)
	if request.method == "POST":
			
		if form.is_valid():
			usuario = request.user

			nombre 				= form.cleaned_data['nombre']
			apellido 			= form.cleaned_data['apellido']
			identificacion 	 	= form.cleaned_data['identificacion']
			correo 				= form.cleaned_data['correo']
			telefono 			= form.cleaned_data['telefono']
			clasificacion 		= form.cleaned_data['Clasificacion_persona']
			declarante 			= form.cleaned_data['declarante']
			pensionado 			= form.cleaned_data['pensionado']
			regimeniva 			= form.cleaned_data['regimen_iva']
			centro 				= form.cleaned_data['centro']
			ingresos 			= form.cleaned_data['ingresos']
			servicios_extra		= form.cleaned_data['servicios_extra']
			objeto_embargo 		= form.cleaned_data['objeto_embargo']
			valor_embargo 		= form.cleaned_data['valor_embargo']

			inst = Contratista.objects.create(

				nombre 					= nombre, 
				apellido 				= apellido, 
				identificacion 			= identificacion, 
				correo 					= correo, 
				telefono				= telefono,
				Clasificacion_persona 	= clasificacion ,
				ingresos				= ingresos,
				declarante 				= declarante, 
				pensionado 				= pensionado,
				regimen_iva 			= regimeniva,
				servicios_extra			= servicios_extra,
				objeto_embargo			= objeto_embargo,
				valor_embargo			= valor_embargo,
				centro 					= centro, 
				user 					= request.user

				)

			inst.save()

			return HttpResponseRedirect("/agregar/cuenta/")
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/datoscontratista.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/datoscontratista.html',locals() )
	

#anyi

def add_comisiones_view(request):
	info = "inicializando"
	if request.method =="POST":
		formucomisiones = add_comisiones_form(request.POST, request.FILES)
		if formucomisiones.is_valid():

			usuario = request.user
			contratista = Contratista.objects.get(user_id = request.user.id)
			contrato = contratista.contrato
			#num_con = Contrato.objects.get( id = contrato)
			#num_contrato = num_con.numero_contrato

			numero_comision = formucomisiones.cleaned_data['numero_comision']
			valor_comision = formucomisiones.cleaned_data['valor_comision']
			valor_retenido = formucomisiones.cleaned_data['valor_retenido']
			total = valor_comision - valor_retenido
 
			comision = Comisiones.objects.create(
				numero_comision = numero_comision, 
				valor_comision = valor_comision,
				valor_retenido = valor_retenido, 
				total= total, 
				user = usuario,
				contrato = contrato)

			comision.save()


			
			return HttpResponseRedirect ('/comilista/page/1/')
	else:
		formucomisiones = add_comisiones_form()
	ctx ={'form': formucomisiones, 'informacion':info}
	return render_to_response('home/comisiones.html', ctx, context_instance = RequestContext(request))

def ver_comisiones_view(request,id_co):
	usuario = request.user
	contratista = Contratista.objects.get(user_id = request.user.id)
	num_con = Contrato.objects.get( contratista_id = contratista.id)
	num_contrato = num_con.numero_contrato
	co = Comisiones.objects.get(id= id_co)
	ctx = {'comisiones':co,
			'contrato':str(num_contrato)}

	return render_to_response('home/ver_comisiones.html',ctx,context_instance = RequestContext(request))

def comilista_view(request, pagina):
	if request.user.is_active:
		suma_comision = 0
		suma_retida_co = 0	

		lista_co = Comisiones.objects.filter(user = request.user.id)
		paginator = Paginator(lista_co, 5)

		for x in lista_co:
			suma_comision = suma_comision +x.valor_comision
			suma_retida_co = suma_retida_co +x.valor_retenido
		try:
			page = int(pagina)
		except:
			page = 1
		try:
			comilista = paginator.page(page)
		except (EmptyPage,InvalidPage):
			comilista = paginator.page(paginator.num_pages)
			
		ctx = {'comilista':comilista,
		'total_comision': suma_comision,
		'total_retenido_co': suma_retida_co}
		return render_to_response ('home/comilista.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/')

def edit_comisiones_view(request, id_co):
	info = ""
	co = Comisiones.objects.get(pk = id_co)
	if request.method =="POST":
		formucomisiones = add_comisiones_form(request.POST, request.FILES, instance= co)
		if formucomisiones.is_valid():

			edit_co = formucomisiones.save(commit = False)
			#t = t + add.horas_laboradas
			edit_co.total=edit_co.valor_comision- edit_co.valor_retenido
			edit_co.save()
			#form.save_m2m()
			info = "Guardado Satisfatoriamente"
			
			return HttpResponseRedirect ('/comisiones/%s'%  edit_co.id)
	else:
		formucomisiones = add_comisiones_form(instance = co)
	ctx ={'form': formucomisiones, 'informacion':info}
	return render_to_response('home/editar_comisiones.html', ctx, context_instance = RequestContext(request))

def del_comisiones_view(request, id_co):
	info = "inicializando"
	try:
		co = Comisiones.objects.get(pk = id_co)
		co.delete()
		info = "Comision Eliminada Correctamente"
		return HttpResponseRedirect('/comilista/page/1/')
	except:
		info = "Comisiones No se puede Eliminar"
		return HttpResponseRedirect('home/comilista/')



def add_honorario_view(request):
	#hono = Contratista.objects.get(id = id_hono)

	info = "inicializando"
	if request.method =="POST":
		formuhonorario = add_honorario_form(request.POST, request.FILES)
		if formuhonorario.is_valid():

			usuario = request.user
			contratista = Contratista.objects.get(user_id = request.user.id)
			contrato = contratista.contrato

			concepto = formuhonorario.cleaned_data['concepto']
			valor_hora = formuhonorario.cleaned_data['valor_hora']
			valor_retenido = formuhonorario.cleaned_data['valor_retenido']
			total = valor_hora - valor_retenido

			
 
			honorarioo = Honorario.objects.create(
				concepto = concepto, 
				valor_hora = valor_hora,
				valor_retenido = valor_retenido, 
				total= total, 
				#contratista = contratista,
				user = usuario
				)

			honorarioo.save()


			
			return HttpResponseRedirect ('/honoralista/page/1/')

			
	else:
		formuhonorario = add_honorario_form()
	ctx ={'form': formuhonorario, 'informacion':info}
	return render_to_response('home/honorario.html', ctx, context_instance = RequestContext(request))




def ver_honorario_view(request,id_hono):

	hono = Honorario.objects.get(id= id_hono)
	ctx = {'honorario':hono}


	return render_to_response('home/ver_honorario.html',ctx,context_instance = RequestContext(request))
	


def honoralista_view(request, pagina):
	suma_hora = 0
	suma_retida = 0
	lista_hono = Honorario.objects.filter(user_id = request.user.id)
	paginator = Paginator(lista_hono, 5)
	for i in lista_hono:
		suma_hora = suma_hora +i.valor_hora
		suma_retida = suma_retida +i.valor_retenido 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		honolista = paginator.page(page)
	except (EmptyPage,InvalidPage):
		honolista = paginator.page(paginator.num_pages)
		
	ctx = {'honolista':honolista,
	'total_hora': suma_hora,
	'total_retenido': suma_retida}
	return render_to_response ('home/honolista.html', ctx, context_instance = RequestContext(request))


def edit_honorario_view(request, id_hono):
	info = ""
	hono = Honorario.objects.get(pk = id_hono)
	if request.method =="POST":
		formuhonorario = add_honorario_form(request.POST, request.FILES, instance= hono)
		if formuhonorario.is_valid():
			edit_hono = formuhonorario.save(commit = False)
			#t = t + add.horas_laboradas
			edit_hono.total=edit_hono.valor_hora- edit_hono.valor_retenido
			edit_hono.save()
			#form.save_m2m()
			info = "Guardado Satisfatoriamente"
			return HttpResponseRedirect ('/honorario/%s'% edit_hono.id)
	else:
		formuhonorario = add_honorario_form(instance = hono)
	ctx ={'form': formuhonorario, 'informacion':info}
	return render_to_response('home/editar_honorario.html', ctx, context_instance = RequestContext(request))

def del_honorario_view(request, id_hono):
	info = "inicializando"
	try:
		hono = Honorario.objects.get(pk = id_hono)
		hono.delete()
		info = "Honorario Eliminado Correctamente"
		return HttpResponseRedirect('/honoralista/page/1/')
	except:
		info = "Honorario No se puede Eliminar"
		return HttpResponseRedirect('home/honoralista/')

def add_actividad_view(request):
	
	if request.user.is_active:
		
		info = "inicializando"
		#t = 0
		if request.method =="POST":
			form = add_actividad_form(request.POST, request.FILES)
			if form.is_valid():

				usuario = request.user

				numero_ficha = form.cleaned_data['numero_ficha']
				descripcion_actividad = form.cleaned_data['descripcion_actividad']
				resultado = form.cleaned_data['resultado']
				horas_laboradas = form.cleaned_data['horas_laboradas']

				actividadd = Actividad.objects.create(numero_ficha = numero_ficha, 
					descripcion_actividad = descripcion_actividad, 
					resultado = resultado, 
					horas_laboradas = horas_laboradas, 
					user = usuario)

				actividadd.save()
				#return redirect('vista_actividades')
				
				
				#t = t + add.horas_laboradas
				#add.save()
				#form.save_m2m()
				#info = "Guardado Satisfatoriamente"
				return HttpResponseRedirect ('/actividades/page/1/')
		else:
			form = add_actividad_form()
		ctx ={'form': form, 'informacion':info}
		return render_to_response('home/actividad.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect ('/login/')

def ver_actividad_view(request,id_acti):
	acti = Actividad.objects.get(id= id_acti)
	ctx = {'actividad':acti}

	return render_to_response('home/ver_actividad.html',ctx,context_instance = RequestContext(request))
	


def actividades_view(request, pagina):
	contratista = Contratista.objects.get(user_id = request.user.id)
	contrato = Contrato.objects.get(contratista_id = contratista.id)
	pagomes = PagoMes.objects.get(contrato = contrato.id)

	fecha_inicio = pagomes.fecha_inicio
	fecha_fin = pagomes.fecha_fin
	
	lista_acti = Actividad.objects.filter(fecha_actividad__range=(fecha_inicio,fecha_fin), user_id = request.user)
	paginator = Paginator(lista_acti, 6)
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		actividades = paginator.page(page)
	except (EmptyPage,InvalidPage):
		actividades = paginator.page(paginator.num_pages)
		
	ctx = {'actividades':actividades}
	return render_to_response ('home/actividades.html', ctx, context_instance = RequestContext(request))

def edit_actividad_view(request, id_acti):
	info = ""
	acti = Actividad.objects.get(pk = id_acti)
	if request.method =="POST":
		form = add_actividad_form(request.POST, request.FILES, instance= acti)
		if form.is_valid():
			edit_acti = form.save(commit = False)
			edit_acti.horas_laboradas =True
			#t = t + add.horas_laboradas
			edit_acti.save()
			#form.save_m2m()
			info = "Guardado Satisfatoriamente"
			return HttpResponseRedirect ('/actividad/%s'% edit_acti.id)
	else:
		form = add_actividad_form(instance = acti)
	ctx ={'form': form, 'informacion':info}
	return render_to_response('home/editar_actividad.html', ctx, context_instance = RequestContext(request))



def del_actividad_view(request, id_acti):
	info = "inicializando"
	try:
		acti = Actividad.objects.get(pk = id_acti)
		acti.delete()
		info = "Actividad Eliminada Correctamente"
		return HttpResponseRedirect('/actividades/page/1')
	except:
		info = "Actividad No se puede Eliminar"
		return HttpResponseRedirect('home/actividades/')

def contrato_view (request):
	contratista = Contratista.objects.get(user_id = request.user.id)
	lista = Contrato.objects.filter(contratista = contratista.id)
	estcontrato = Contrato.objects.get(contratista = contratista.id)
	estado = estcontrato.estado
	print estado
	ctx = {'lista':lista}
	return render(request, 'home/contrato.html',locals() )


	#return render_to_response ( 'home/contrato.html', ctx, context_instance = RequestContext(request))

def edit_contrato_view(request, id_cont):
	info ="" 
	cont = Contrato.objects.get(pk = id_cont)
	if request.method == "POST":
		formulario = add_contrato_form (request.POST, request.FILES, instance= cont)
		if formulario.is_valid():
			edit_contrato = formulario.save(commit = False)
			formulario.save_m2m()

			edit_contrato.save()
			info ="Guardado Satisfactoriamente"
			return HttpResponseRedirect('/contrato')
	else:
		formulario = add_contrato_form(instance = cont )
	ctx ={'form' :formulario,'informacion': info}
	return render_to_response('home/add_contrato.html', ctx,context_instance = RequestContext(request))
	
def add_contrato_view(request):
	info = "inicializando"
	vista = 'Contrato'
	if request.method == "POST":

		formulario = ContratoForm(request.POST, request.FILES)
		if formulario.is_valid():
			usuario = request.user.id
			contratista = Contratista.objects.get(user = usuario)

			num_contrato 		= formulario.cleaned_data['numero_contrato']
			fechainicio 		= formulario.cleaned_data['fecha_inicio']
			fechafin 			= formulario.cleaned_data['fecha_fin']
			valortotal 			= formulario.cleaned_data['valor_total']
			compromiso 			= formulario.cleaned_data['compromiso_siif']
			forma_estimacion 	= formulario.cleaned_data['forma_estimacion']
			ordenador_pago		= formulario.cleaned_data['ordenador_pago']
			supervisor 			= formulario.cleaned_data['supervisor']
			valor_mensual		= formulario.cleaned_data['valor_mensual']
			meses 				= fechafin.month - fechainicio.month
			total_horas_contrato = meses*160

			contrato = Contrato.objects.create(
				numero_contrato = num_contrato, 
				fecha_inicio= fechainicio, 
				fecha_fin = fechafin ,
				valor_total = valortotal ,
				compromiso_siif = compromiso, 
				forma_estimacion = forma_estimacion ,
				valor_mensual =   valor_mensual,
				total_horas_contrato = total_horas_contrato,
				ordenador_pago = ordenador_pago, 
				supervisor=supervisor, contratista = contratista)
					
			contrato.save()

			pago = PagoMes.objects.create(
				valor_mes = valor_mensual,
				total_dias_liquidacion = 0,
				fecha_inicio = fechainicio,
				fecha_fin = fechafin,
				saldo_anterior = valortotal,
				saldo_nuevo = valortotal,
				total_pagar = 0,
				numero_pago = 0,
				contrato = contrato
				)
			pago.save()
			return HttpResponseRedirect('/contrato/')
	else:
		formulario = ContratoForm()

	ctx = {'form':formulario, 'informacion':info, 'vista':vista}
	return render_to_response ('home/add_contrato.html', ctx,context_instance = RequestContext(request))


def add_ordenador_view(request):

	form = Ordenador_pagoForm(request.POST or None)
	if request.method == "POST":
			
		if form.is_valid():
			instance = form.save(commit=False)
			instance.save()
			context = {
				"form": form
			} 
			return HttpResponseRedirect('/gestion/')
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/agregarOrdenadorPago.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/agregarOrdenadorPago.html',locals() )




def agregarRegional_View(request):

	form = Registrar_regionalForm(request.POST or None)
	if request.method == "POST":
			
		if form.is_valid():
			instance = form.save(commit=False)
			instance.save()
			#	print  request.POST.get("codigo")
			#	nombre = request.POST.get("nombre_regional")
			# 	Post.objects.create(codigo=codigo)
			context = {
				"form": form
			} 
			return render(request, 'home/index.html',locals() )
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/registrar_regional.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/registrar_regional.html',locals() )

def agregarCentro_View(request):
	form = Registrar_centroForm(request.POST or None)
	if request.method == "POST":

		if form.is_valid():
			instance = form.save(commit=False)
			instance.save()
		#	if request.method == "POST":
		#	print  request.POST.get("codigo")
		#	nombre = request.POST.get("nombre_regional")
		# 	Post.objects.create(codigo=codigo)
		context = {
			"form": form
		} 
		return render(request, 'home/index.html',locals() )
		

	else:
		context = {
				"form": form
			} 
		return render(request, 'home/registrar_centro.html',locals() )

def agregarBanco_View(request):

	form = Registrar_Banco(request.POST or None)
	if request.method == "POST":
		if form.is_valid():
			instance = form.save(commit=False)
			instance.save()
			context = {
				"form": form
			} 
			return HttpResponseRedirect('/gestion/')
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/agregarbanco.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/agregarbanco.html',locals() )

def agregarCuenta_View(request):

	form = Registrar_Cuenta(request.POST or None)
	if request.method == "POST":
		if form.is_valid():
			contratista = Contratista.objects.get(user_id = request.user.id)
			num_cuenta	 		= form.cleaned_data['num_cuenta']
			tipo_cuenta 		= form.cleaned_data['tipo_cuenta']
			banco_cuenta		= form.cleaned_data['banco_cuenta']
			

			cuenta = Cuenta.objects.create(
				num_cuenta = num_cuenta, 
				tipo_cuenta= tipo_cuenta, 
				banco_cuenta = banco_cuenta,
				contratista = contratista
				 )
			cuenta.save()
			return HttpResponseRedirect('/ver/contratista/')
		else:
			context = {
				"form": form
			} 
			return render(request, 'home/agregarCuenta.html',locals() )
	else:
		context = {
				"form": form
			} 
		return render(request, 'home/agregarCuenta.html',locals() )

		

def planilla_view(request):
	return render_to_response('home/planilla.html',context_instance = RequestContext(request))

def single_planilla_view(request, id_plani):
	planilla_view = planilla.objects.filter(status = True)
	palni = planilla.objects.get(id = id_plani)	
	ctx = { 'planilla':planilla ,'planilla':plani}
	return render_to_response('home/single_planilla.html',ctx,context_instance = RequestContext(request))	

#chilito

def demo_view(request):
	try:
		contrato = Contrato.objects.get(id = 1)#cambiar por el contratista que esta logueado , contratists__user__id, contratista__identificacion
		calc_ibc = contrato.valor_mensual * 0.40

		print "si lo hago" , calc_ibc
	except:
		pass
	if request.method == "POST":
		formulario = aportes(request.POST)
		if formulario.is_valid():
			apo = formulario.save(commit = False)
			apo.ibc_aporte = int(calc_ibc) 
			apo.pension_aporte = int(calc_pension) 
			apo.save()
	else:
		pass
	formulario = aportes()
	ctx = {'form':formulario}
	return render_to_response('home/aportes.html',ctx,context_instance=RequestContext(request))
	
def aportes_view(request):
	
	if request.user.is_active:
		
		info = "inicializando"
		if request.method =="POST":
			form = Aportes_Form(request.POST, request.FILES)
			if form.is_valid():

				usuario = request.user

				numero_plantilla = form.cleaned_data['numero_plantilla']
				tipoRiesgo_arl = form.cleaned_data['tipoRiesgo_arl']
				aporte_vpension = form.cleaned_data['aporte_vpension']
				aporte_vafc = form.cleaned_data['aporte_vafc']
				prestamo_vivienda = form.cleaned_data['prestamo_vivienda']
				medicina_prepagada = form.cleaned_data['medicina_prepagada']

				#contratista = Contratista.objects.get(user_id = request.user.id)
				#contrato = Contrato.objects.get(contratista_id = contratista.id)
				#calc_ibc = (contrato.valor_mensual * 0.04)
				#calc_salud = (contrato.valor_mensual * 12.5)
				#calc_pension = (contrato.valor_mensual  * 0.16)
				#calc_arl = (contrato.valor_mensual * 0.05)

				aportes = Aporte.objects.create(

					numero_plantilla = numero_plantilla, 
					tipoRiesgo_arl = tipoRiesgo_arl, 
					aporte_vpension = aporte_vpension, 
					aporte_vafc = aporte_vafc,
					prestamo_vivienda = prestamo_vivienda,
					medicina_prepagada = medicina_prepagada,
					ibc_aporte = 0,
					salud_aporte = 0,
					pension_aporte = 0,
					solidaridad_aporte = 0,
					arl_aporte = 0,
					valor_arl = 0,
					user = usuario
					)

				aportes.save()
				return HttpResponseRedirect ('/')
		else:
			form = Aportes_Form()
		ctx ={'form': form, 'informacion':info}
		return render_to_response('home/aportes.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect ('/login/')

def contratista_view(request):
	form =  ContratistaForm()
	if request.method == "POST":
		form = ContratistaForm(request.POST)
		if form.is_valid():
			usuario = form.cleaned_data['username']
			email = form.cleaned_data['email']
			password_one = form.cleaned_data['password_one']
			password_two = form.cleaned_data['password_two']
			u = User.objects.create_user(username=usuario,email=email,password=password_one)
			u.save()
			send_mail(
	   			'Registro',
				'su usuario es:'+usuario+ '' 'su password es:'+password_one,
				'planipagorestaurar@gmail.com',
				[''+email],
				fail_silently=False,
	   		)	
			return HttpResponseRedirect ('/contratistas/')
		else:
			ctx = {'form':form}
			return render_to_response('home/registrarCon.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form}
	return render_to_response('home/registrarCon.html',ctx,context_instance=RequestContext(request))

def pagoMes_view(request):
	saldo_Anterior=0
	nuevo_saldo=0
	valor_contrato=0

	try:
		contratista = Contratista.objects.get(user_id = request.user.id)
		contrato = Contrato.objects.get(contratista_id = contratista.id)
		calc_ibc = (contrato.valor_mensual * 0.04)
		calc_salud = (contrato.valor_mensual * 12.5)
		calc_pension = (contrato.valor_mensual  * 0.16)
		calc_arl = (contrato.valor_mensual * 0.05)
	except:
		pass
	form = PagoMesForm(request.POST or None)
	if request.method == "POST":
		if form.is_valid():
			contratista 	= Contratista.objects.get(user_id = request.user.id)
			contrato 		= Contrato.objects.get(contratista_id = contratista.id)
			pagomes 		= PagoMes.objects.get(contrato_id = contrato.id)

			dias 					= form.cleaned_data['total_dias_liquidacion']
			fechainicio 			= form.cleaned_data['fecha_inicio']
			fechafin 				= form.cleaned_data['fecha_fin']
			numero_delpago 			= form.cleaned_data['numero_pago']

			valor_dia 				= contrato.valor_mensual/30
			total_pagar 			= dias * valor_dia

			pagomes.total_dias_liquidacion  = dias
			pagomes.fecha_inicio 	 		= fechainicio
			pagomes.fecha_fin 		 		= fechafin
			pagomes.saldo_anterior   		= pagomes.saldo_nuevo
			pagomes.saldo_nuevo				= pagomes.saldo_anterior - total_pagar
			pagomes.numero_pago 	 		= numero_delpago
			pagomes.total_pagar		 		= total_pagar
			pagomes.save()

			valor = pagomes.total_pagar + pagomes.saldo_anterior


			#pago = PagoMes.objects.create(
			#	valor_mes = contrato.valor_mensual,
			#	total_dias_liquidacion = dias,
			#	fecha_inicio = fechainicio,
			#	fecha_fin = fechafin,
			#	saldo_anterior = saldo_anterior,
			#	saldo_nuevo = saldo_nuevo,
			#	total_pagar = total_pagar,
			#	numero_pago = numero_delpago,
			#	contrato = contrato
			#	)
			#pago.save()
			

		context = {
			"form": form
		} 
		return HttpResponseRedirect('/comilista/page/1/')
		
	else:
			ctx = {'form':form, 'valor_mes':contrato.valor_mensual}
			return render_to_response('home/pagomes.html',ctx,context_instance=RequestContext(request))

	ctx = {'form':form, 'valor_mes':contrato.valor_mensual}
	return render_to_response('home/pagomes.html',ctx,context_instance=RequestContext(request))


#ELIANA

def add_contratista_view(request):
	info = "inicializando"
	if request.method == "POST":
		formulario = add_contratista_form(request.POST, request.FILES)
		if formulario.is_valid():
			add = formulario.save(commit = False)
			add.save() # guarda la informacion
			formulario.save_m2m() # guarda las relaciones ManyToMany
			info = "Guardado Satisfactoriamente"
			return HttpResponseRedirect ('/contratista/%s' %add.id)
	else:
		formulario = add_contratista_form()
	ctx = {'form':formulario, 'informacion':info}
	return render_to_response('home/add_contratista.html', ctx,context_instance = RequestContext(request))

def single_contratista_view(request, id_contratista):
	inge = Contratista.objects.get(id = id_contratista)
	ctx = {'contratista':inge}
	return render_to_response('home/single_contratista.html',ctx,context_instance = RequestContext(request))

def contratistas_view(request):
	if request.user.is_staff:
	
		lista_inge = Contratista.objects.filter()
		ctx = {'contratistas':lista_inge}
		return render_to_response ('home/listacontratistas.html', ctx, context_instance = RequestContext(request))
	else:
		return HttpResponseRedirect('/')

def edit_contratista_view(request, id_contratista):
	info = ""
	inge = Contratista.objects.get(pk = id_contratista)
	if request.method == "POST":
		formulario = add_contratista_form(request.POST, request.FILES, instance= inge)
		if formulario.is_valid():
			edit_inge = formulario.save(commit = False)
			edit_inge.status = True
			formulario.save_m2m()
			edit_inge.save()
			info = "Guardado Satisfactoriamente"
			return HttpResponseRedirect('/contratista/%s'% edit_inge.id)
	else:
		formulario = add_contratista_form(instance = inge)
	ctx = {'form':formulario, 'informacion':info}
	return render_to_response('home/edit_contratista.html', ctx,context_instance = RequestContext(request))

#ELIANA
def estado_contrato_view (request, id_contrato):
	#if request.user.is_authenticated and request.user.is_staff:

	info = ""
	#inge = Contratista.objects.get(pk = id_contratista)
	inge = Contrato.objects.get(pk = id_contrato)
	if request.method == "POST":
		formulario = estado_contrato_form(request.POST,request.FILES, instance= inge) #FILES para agregar imagenes
		if formulario.is_valid():
			contrato_inge = formulario.save(commit = False)
			#formulario.save_m2m()
			contrato_inge.save()
			info = "Guardado satisfactoriamente"
			return HttpResponseRedirect ('/contratista/%s'% contrato_inge.id)		
	elif request.method =="GET":
		formulario = estado_contrato_form(instance = inge)

	ctx = {'form':formulario, 'informacion':info}
	return render_to_response('home/estado_del_contrato.html', ctx,context_instance = RequestContext(request))

def listacontratista_view (request):
	lista = Contratista.objects.filter(user_id = request.user.id)
	ctx = {'lista':lista}
	return render_to_response ( 'home/contratista_view.html', ctx, context_instance = RequestContext(request))



def listaaportes_view (request):
	listaaportes = Aporte.objects.filter(user_id = request.user.id)
	ctx = {'listaaportes':listaaportes}
	return render_to_response ( 'home/vista_aportes.html', ctx, context_instance = RequestContext(request))


def edit_aporte_view(request, id_aporte):
	info = ""
	aporte = Aporte.objects.get(pk = id_aporte)
	if request.method == "POST":
		formulario = Aportes_Form(request.POST, request.FILES, instance= aporte)
		if formulario.is_valid():
			edit_aporte = formulario.save(commit = False)
			edit_aporte.save()
			info = "Guardado Satisfactoriamente"
			return HttpResponseRedirect('/ver/aportes/')
	else:
		formulario = Aportes_Form(instance = aporte)
	ctx = {'form':formulario, 'informacion':info}
	return render_to_response('home/edit_aporte.html', ctx,context_instance = RequestContext(request))

def edit_numeroplantilla_view(request):
	info = ""
	if request.method == "POST":
		formulario = Aportes_Form2(request.POST, request.FILES)
		if formulario.is_valid():
			numero_plantilla  				= formulario.cleaned_data['numero_plantilla']
			aportes 						= Aporte.objects.get (user_id = request.user.id)
			aportes.numero_plantilla 		= numero_plantilla
			aportes.save()
			info = "Guardado Satisfactoriamente"
			return HttpResponseRedirect('/actividades/page/1/')
	else:
		formulario = Aportes_Form2()
	ctx = {'form':formulario, 'informacion':info}
	return render_to_response('home/edit_numeroplantilla.html', ctx,context_instance = RequestContext(request))



def mostrarTabla(request):

	#contratistas = Contratista.objects.get(user = request.user) ASI DEBE IR

	contratista  = 	Contratista.objects.get(user_id = request.user.id)
	

	contratos 	= 	Contrato.objects.get(contratista_id = contratista.id)
	supervisores = 	Supervisor.objects.get(id = contratos.supervisor_id)
	pagomes 	= 	PagoMes.objects.get(contrato = contratos.id)

	fecha_inicio	 = 	 pagomes.fecha_inicio
	fecha_fin 		 = 	 pagomes.fecha_fin
	
	actividades  = 	Actividad.objects.filter(fecha_actividad__range=(fecha_inicio,fecha_fin), user_id =request.user)
	cuenta 	   	 = 	Cuenta.objects.get(contratista_id = contratista.id)

	banco 	   	 = 	Banco.objects.get(id = cuenta.banco_cuenta_id)
	contrato   	 = 	Contrato.objects.filter(contratista_id = contratista.id)
	pago_mes   	 = 	PagoMes.objects.filter(contrato = contratista.contrato )
	comisiones	 = 	Comisiones.objects.filter(contrato = contratos)
	centro 		 = 	Centro.objects.filter(contratista = contratista)

	return render(request,'home/tabla.html',locals())



	
def llenarDatos_view(request):
	
	#cONSULTAS GENERALES
	contratista  = 	Contratista.objects.get(user_id = request.user.id)
	contrato = Contrato.objects.get(contratista_id = contratista.id)
	cuenta 	   	 = 	Cuenta.objects.get(contratista_id = contratista.id)
	pagomes 	= 	PagoMes.objects.get(contrato = contrato.id)

	
	template_name = 'home/descargar.html'
	response = render_to_response('home/descargar.html')
	#wb = load_workbook(filename = 'planilla/media/excel/DATOSPRUEBA.xlsx')
	wb = load_workbook(filename = 'planilla/media/excel/PLANILLAVERSION.xlsx')
	nombre = contratista.nombre + " " + contratista.apellido 
	#wb = Workbook()

	#dest_filename = 'planilla/media/excel/DATOSPRUEBA.xlsx'
	#dest_filename = 'planilla/media/excel/PLANILLAVERSION.xlsx'


	wb.get_sheet_names()
	#hoja=wb.get_sheet_by_name('DATOS')

	hoja=wb.get_sheet_by_name('DATOS DE ENTRADA.')
	#DATOS DEL CONTRATISTA
	hoja['B5']=contratista.centro.regional_centro.codigo
	hoja['B6']=contratista.centro.codigo
	hoja['B7']=contratista.centro.codigo #AQUI VA LA FECHA DE PLANILLA
	hoja['B8']=nombre
	hoja['B9']=contratista.identificacion
	hoja['B10']=contratista.telefono
	hoja['B11']=contratista.correo
	hoja['B12']=contratista.centro.regional_centro.nombre_regional
	hoja['B13']=contratista.centro.nombre_centro
	hoja['B14']=cuenta.banco_cuenta.nombre_banco
	hoja['B15']=cuenta.tipo_cuenta
	hoja['B16']=cuenta.num_cuenta
	hoja['B17']=contratista.Clasificacion_persona
	hoja['B18']=contratista.ingresos
	hoja['B19']=contratista.regimen_iva
	hoja['B20']=contratista.servicios_extra
	hoja['B21']=contratista.declarante
	hoja['B22']=contratista.pensionado
	hoja['B23']=contratista.objeto_embargo
	hoja['B24']=contratista.valor_embargo

	#DATOS DEL SUPERVISOR
	nombresupervisor = contrato.supervisor.nombre + " " + contrato.supervisor.apellido
	hoja['B27']=nombresupervisor
	hoja['B29']=contrato.supervisor.correo

	#DATOS DEL SUPERVISOR
	nombreordenadorpago = contrato.ordenador_pago.nombre + " " + contrato.ordenador_pago.apellido
	hoja['B32']=nombreordenadorpago

	#DATOS DEL CONTRATO
	hoja['B37']=contrato.numero_contrato
	hoja['B38']=contrato.fecha_inicio
	hoja['B39']=contrato.fecha_fin
	hoja['B40']=contrato.valor_total
	hoja['B41']=contrato.compromiso_siif
	hoja['B42']=contrato.forma_estimacion

	#DATOS DEL PAGO
	hoja['B45']=pagomes.fecha_inicio
	hoja['B46']=pagomes.fecha_fin
	hoja['B47']=pagomes.numero_pago
	hoja['B48']=pagomes.saldo_anterior
	hoja['B49']=pagomes.saldo_nuevo

	#PAGO POR PERIODO
	hoja['B57']=contrato.valor_mensual
	hoja['B58']=pagomes.total_dias_liquidacion
	hoja['B59']=pagomes.total_pagar

	response = HttpResponse(content=save_virtual_workbook(wb), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename='+nombre+'.xlsx'
	return response
	#response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
	#return response


	